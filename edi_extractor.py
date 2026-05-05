#!/usr/bin/env python3
"""
EDI Field Extractor - Core extraction logic without GUI dependencies
Extracts EDI field descriptions from PDF specification documents
"""

from pathlib import Path
import re
import csv
from typing import List, Tuple
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class EDIFieldExtractor:
    """Extracts EDI field descriptions and codes from PDF"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = Path(pdf_path)
        self.text_content = ""
        self.field_mappings: List[Tuple[str, str]] = []
        self.segment_groups = {}  # Store segment groupings
        self.code_lists = {}  # Store code lists for fields that need expansion
        
    def extract_text(self, start_page: int = 2) -> str:
        """Extract text from PDF file, skipping preliminary sections"""
        try:
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_parts = []
                total_pages = len(pdf_reader.pages)
                
                actual_start = self._find_content_start(pdf_reader, start_page)
                
                for page_num in range(actual_start, total_pages):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    text_parts.append(page_text)
                
                self.text_content = "\n".join(text_parts)
                return self.text_content
        except Exception as e:
            raise Exception(f"Error extracting PDF text: {str(e)}")
    
    def _find_content_start(self, pdf_reader, default_start: int = 2) -> int:
        """Find the actual start of content by looking for 'Data Element Summary'"""
        for page_num in range(min(10, len(pdf_reader.pages))):
            page_text = pdf_reader.pages[page_num].extract_text()
            if 'Data Element Summary' in page_text or 'SEGMENT:' in page_text:
                return page_num
        return default_start
    
    def parse_field_mappings(self) -> List[Tuple[str, str]]:
        """Parse field descriptions and codes from extracted text"""
        if not self.text_content:
            raise ValueError("No text content available. Call extract_text() first.")
        
        # First, extract code lists for fields that need expansion
        self._extract_code_lists()
        
        # Parse field mappings
        self.field_mappings = self._parse_data_element_summary()
        
        # Expand N1 segments if code lists are available
        if self.code_lists:
            self.field_mappings = self._expand_n1_segments(self.field_mappings)
        
        return self.field_mappings
    
    def _extract_code_lists(self):
        """Extract code lists for fields that need expansion (like N101)"""
        # Pattern to match code list sections
        code_list_pattern = r'Code\s+Name\s*\n((?:[A-Z0-9]{1,3}\s+[^\n]+\n)+)'
        
        # Look for N101 code list specifically
        n101_section = re.search(r'N101.*?Code\s+Name\s*\n((?:[A-Z0-9]{1,3}\s+[^\n]+\n)+)', 
                                 self.text_content, re.DOTALL)
        
        if n101_section:
            codes_text = n101_section.group(1)
            codes = []
            for line in codes_text.split('\n'):
                line = line.strip()
                if line and len(line) > 3:
                    # Split on first whitespace to separate code from description
                    parts = line.split(None, 1)
                    if len(parts) == 2:
                        code, description = parts
                        codes.append((code, description))
            
            if codes:
                self.code_lists['N101'] = codes
    
    def _parse_data_element_summary(self) -> List[Tuple[str, str]]:
        """Parse Data Element Summary sections with improved flexibility"""
        field_mappings = []
        
        # Try multiple patterns to match different PDF formats
        patterns = [
            # Pattern 1: Standard "Data Element Summary" format
            r'Data Element Summary\s*\n.*?Ref\.\s+Des\.\s+Element\s+Name\s+Attributes\s*\n(.*?)(?=Data Element Summary|\Z)',
            # Pattern 2: Alternative format with "SEGMENT:"
            r'SEGMENT:\s*([A-Z0-9]+).*?\n(.*?)(?=SEGMENT:|\Z)',
            # Pattern 3: Simple table format
            r'(?:Ref|Element|Field).*?(?:Des|Code|ID).*?(?:Name|Description).*?\n(.*?)(?=\n\n|\Z)',
        ]
        
        for pattern in patterns:
            summaries = list(re.finditer(pattern, self.text_content, re.DOTALL | re.IGNORECASE))
            if summaries:
                for summary in summaries:
                    content = summary.group(1) if len(summary.groups()) == 1 else summary.group(2)
                    
                    # Parse each line in the summary
                    lines = content.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line or line.startswith('---') or line.startswith('===') or len(line) < 5:
                            continue
                        
                        # Try multiple extraction patterns
                        code = None
                        description = None
                        
                        # Pattern 1: Standard format (code + description)
                        match = re.match(r'^([A-Z0-9]{2,6})\s+(.+?)(?:\s+[MO]\s+|\s+[A-Z]{1,3}\d+|\s*$)', line)
                        if match:
                            code = match.group(1).strip()
                            description = match.group(2).strip()
                        else:
                            # Pattern 2: Alternative format
                            match = re.match(r'^([A-Z0-9]{2,6})\s*[-:]\s*(.+?)(?:\s+[MO]\s+|\s*$)', line)
                            if match:
                                code = match.group(1).strip()
                                description = match.group(2).strip()
                            else:
                                # Pattern 3: Tab or multiple space separated
                                parts = re.split(r'\s{2,}|\t', line, maxsplit=1)
                                if len(parts) == 2 and re.match(r'^[A-Z0-9]{2,6}$', parts[0]):
                                    code = parts[0].strip()
                                    description = parts[1].strip()
                        
                        if code and description:
                            # Clean up description
                            description = re.sub(r'\s+', ' ', description)
                            description = description.replace('  ', ' ')
                            # Remove common suffixes
                            description = re.sub(r'\s+[MO]\s*$', '', description)
                            description = re.sub(r'\s+[A-Z]{1,3}\d+\s*$', '', description)
                            
                            if description and len(description) > 3:
                                field_mappings.append((description, code))
                
                # If we found mappings with this pattern, stop trying others
                if field_mappings:
                    break
        
        return field_mappings
    
    def _expand_n1_segments(self, field_mappings: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
        """Expand N1 segments based on code lists"""
        expanded_mappings = []
        n1_codes = self.code_lists.get('N101', [])
        
        for desc, code in field_mappings:
            # Check if this is an N1 segment field
            if code.startswith('N1') and n1_codes:
                # For N101, create separate columns for each entity identifier code
                if code == 'N101':
                    for entity_code, entity_desc in n1_codes:
                        expanded_desc = f"{desc} - {entity_code} ({entity_desc})"
                        expanded_mappings.append((expanded_desc, f"{code}_{entity_code}"))
                else:
                    # For other N1 fields, create columns for each entity type
                    for entity_code, entity_desc in n1_codes:
                        expanded_desc = f"{desc} - {entity_code}"
                        expanded_mappings.append((expanded_desc, f"{code}_{entity_code}"))
            else:
                expanded_mappings.append((desc, code))
        
        return expanded_mappings
    
    def generate_excel_file(self, field_mappings: List[Tuple[str, str]] = None):
        """Generate Excel file with field mappings"""
        if field_mappings is None:
            field_mappings = self.field_mappings
        
        if not field_mappings:
            raise ValueError("No field mappings available")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "EDI Fields"
        
        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers (descriptions in row 1, codes in row 2)
        for col_idx, (desc, code) in enumerate(field_mappings, start=1):
            # Description header
            cell = ws.cell(row=1, column=col_idx, value=desc)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Code header
            cell = ws.cell(row=2, column=col_idx, value=code)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col_idx, (desc, code) in enumerate(field_mappings, start=1):
            column_letter = get_column_letter(col_idx)
            # Set width based on description length, with min and max limits
            desc_length = len(desc)
            adjusted_width = min(max(desc_length * 0.8, 15), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 25
        
        # Save files
        output_base = str(self.pdf_path).rsplit('.', 1)[0] + '_output'
        excel_file = output_base + '.xlsx'
        csv_file = output_base + '.csv'
        
        # Save Excel
        wb.save(excel_file)
        
        # Save CSV
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Write descriptions
            writer.writerow([desc for desc, _ in field_mappings])
            # Write codes
            writer.writerow([code for _, code in field_mappings])
        
        return excel_file, csv_file

# Made with Bob
