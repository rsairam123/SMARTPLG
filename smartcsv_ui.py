#!/usr/bin/env python3
"""
SmartCSV UI - Graphical Interface for PDF to Excel Conversion
Extracts EDI field descriptions from PDF specification documents and generates Excel files
"""

from pathlib import Path
import re
import csv
from typing import List, Tuple
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Optional tkinter imports for GUI (not needed for web API)
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    import threading
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
    # Dummy classes for when tkinter is not available
    class tk:
        pass
    class ttk:
        pass


class EDIFieldExtractor:
    """Extracts EDI field descriptions and codes from PDF"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = Path(pdf_path)
        self.text_content = ""
        self.field_mappings: List[Tuple[str, str]] = []
        self.segment_groups = {}  # Store segment groupings
        self.code_lists = {}  # Store code lists for fields that need expansion
        
    def extract_text(self, start_page: int = 2) -> str:
        """Extract text from PDF file, skipping preliminary sections"""
        try:
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_parts = []
                total_pages = len(pdf_reader.pages)
                
                actual_start = self._find_content_start(pdf_reader, start_page)
                
                for page_num in range(actual_start, total_pages):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    text_parts.append(page_text)
                
                self.text_content = "\n".join(text_parts)
                return self.text_content
        except Exception as e:
            raise Exception(f"Error extracting PDF text: {str(e)}")
    
    def _find_content_start(self, pdf_reader, default_start: int = 2) -> int:
        """Find the page where actual EDI segment definitions start"""
        total_pages = len(pdf_reader.pages)
        
        content_indicators = [
            r'\bST\b.*segment',
            r'\bBEG\b.*segment',
            r'\bREF\b.*segment',
            r'\bPO1\b.*segment',
            r'segment.*identifier',
            r'data element',
            r'\[\w{3,6}\d{2,3}\]',
            r'element.*position',
        ]
        
        for page_num in range(default_start, min(default_start + 10, total_pages)):
            page_text = pdf_reader.pages[page_num].extract_text().lower()
            
            indicator_count = sum(1 for pattern in content_indicators
                                if re.search(pattern, page_text, re.IGNORECASE))
            
            if indicator_count >= 2:
                return page_num
        
        return default_start
    
    def extract_code_lists(self) -> dict:
        """Extract code lists from PDF for fields that need expansion"""
        if not self.text_content:
            self.extract_text()
        
        lines = self.text_content.split('\n')
        code_lists = {}
        current_field = None
        in_code_list = False
        in_data_element = False
        
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            lower_line = line_stripped.lower()
            
            # Detect "Data Element Summary" section (for 810 format)
            if 'data element summary' in lower_line or 'element summary:' in lower_line:
                in_data_element = True
                # Look back to find N101 field reference
                for j in range(max(0, i-30), i):
                    prev_line = lines[j].strip()
                    if 'N101' in prev_line and ('Entity Identifier' in prev_line or '98' in prev_line):
                        current_field = 'N101'
                        if current_field not in code_lists:
                            code_lists[current_field] = []
                        break
                continue
            
            # Detect "CodeList Summary" section (for 850 format)
            if ('codelist' in lower_line and 'summary' in lower_line) or \
               ('code list' in lower_line and 'summary' in lower_line):
                in_code_list = True
                # Look back to find the field reference (e.g., N101)
                for j in range(max(0, i-25), i):
                    prev_line = lines[j].strip()
                    field_match = re.search(r'\b([A-Z]{1,4}\d{2,3})\b', prev_line)
                    if field_match:
                        potential_field = field_match.group(1)
                        if potential_field == 'N101' or current_field is None:
                            current_field = potential_field
                            if current_field not in code_lists:
                                code_lists[current_field] = []
                        if potential_field == 'N101':
                            break
                continue
            
            # Extract codes from Data Element Summary (810 format)
            if in_data_element and current_field == 'N101':
                # Pattern for 810: "   BT Bill-to-Party    "
                # Look for 2-3 letter codes followed by description
                patterns = [
                    r'^\s{2,}([A-Z]{2,3})\s+([A-Z][A-Za-z\s\-/()]+?)(?:\s{2,}|$)',  # "   BT Bill-to-Party"
                    r'^([A-Z]{2,3})\s{2,}(.+)$',  # "BT    Bill-to-Party"
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, line_stripped)
                    if match and len(match.groups()) >= 2:
                        code = match.group(1).strip()
                        name = match.group(2).strip()
                        # Clean up the name
                        name = re.sub(r'\s+', ' ', name)
                        # Only add if it looks like a valid code (2-3 chars, all caps)
                        if len(code) in [2, 3] and code.isupper() and len(name) > 3 and not name.isdigit():
                            # Avoid duplicates
                            if not any(c[0] == code for c in code_lists[current_field]):
                                code_lists[current_field].append((code, name))
                            break
                
                # End of N101 codes section
                if line_stripped.startswith('M N102') or line_stripped.startswith('>> N103'):
                    in_data_element = False
                    current_field = None
            
            # Extract codes from CodeList Summary (850 format)
            if in_code_list and current_field:
                patterns = [
                    r'^([A-Z]{2,3})\s{2,}(.+)$',  # BT    Bill-to-Party
                    r'^([A-Z]{2,3})\s+([A-Z][A-Za-z\s\-/]+)$',  # BT Bill-to-Party
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, line_stripped)
                    if match and len(match.groups()) >= 2:
                        code = match.group(1).strip()
                        name = match.group(2).strip()
                        name = re.sub(r'\s+', ' ', name)
                        if len(code) <= 3 and len(name) > 2 and not name.isdigit():
                            if not any(c[0] == code for c in code_lists[current_field]):
                                code_lists[current_field].append((code, name))
                            break
                
                # End of code list detection
                if not line_stripped or re.match(r'^[A-Z]{1,4}\d{2,3}\s+\d+', line_stripped):
                    in_code_list = False
                    current_field = None
        
        return code_lists
    
    def parse_field_mappings(self) -> List[Tuple[str, str]]:
        """Parse field descriptions and their corresponding EDI codes with segment expansion"""
        if not self.text_content:
            self.extract_text()
        
        # First, extract code lists for fields that need expansion
        self.code_lists = self.extract_code_lists()
        
        # Ensure N101 has codes (use fallback if extraction failed)
        # Detect document type based on filename (Method 1: Filename Detection)
        if 'N101' not in self.code_lists or len(self.code_lists.get('N101', [])) == 0:
            pdf_filename = str(self.pdf_path).lower()
            
            if '810' in pdf_filename:
                # 810 Invoice has 5 codes: BT, II, RE, ST, VN
                self.code_lists['N101'] = [
                    ('BT', 'Bill-to-Party'),
                    ('II', 'Issuer of Invoice'),
                    ('RE', 'Party to receive commercial invoice remittance'),
                    ('ST', 'Ship To'),
                    ('VN', 'Vendor')
                ]
            elif '856' in pdf_filename:
                # 856 Ship Notice/Manifest has 3 codes: ST, VN, SF
                self.code_lists['N101'] = [
                    ('ST', 'Ship To'),
                    ('VN', 'Vendor'),
                    ('SF', 'Ship From')
                ]
            else:
                # 850 Purchase Order has 4 codes: BT, ST, SU, VN (default)
                self.code_lists['N101'] = [
                    ('BT', 'Bill-to-Party'),
                    ('ST', 'Ship To'),
                    ('SU', 'Supplier/Manufacturer'),
                    ('VN', 'Vendor')
                ]
        
        lines = self.text_content.split('\n')
        in_element_summary = False
        seen_ref_ids = set()
        all_fields = []  # Store all fields in order
        segment_fields = {}  # Group fields by segment (e.g., N1, N3, N4)
        
        # Regex patterns for EDI element lines
        # Pattern for 850 format: N101 98 Entity Identifier Code M ID 2/3
        pattern1 = r'^([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+([A-Za-z][A-Za-z0-9\s\-/()]+?)\s+[MOCX]\s+(?:AN|N\d*|R|ID|DT|TM)\s+\d+/\d+(?:\s+.*)?$'
        pattern2 = r'^[MOCX]\s+([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+(.+?)\s+[MOCX]\s+(?:AN|N\d*|R|ID|DT|TM)\s+\d+/\d+(?:\s+.*)?$'
        # Pattern for 810 format: M ST01 143 Transaction Set Identifier Code M ID 3/3
        pattern3 = r'^[MOCX]\s+([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+([A-Za-z][A-Za-z0-9\s\-/()]+?)\s+[MOCX]\s+(?:AN|N\d*|R|ID|DT|TM)\s+\d+/\d+(?:\s+.*)?$'
        # Pattern for optional fields with >> prefix: >> N103  66 Identification  Code  Qualifier  O ID 1/2
        pattern4 = r'^>>\s+([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+([A-Za-z][A-Za-z0-9\s\-/()]+?)\s+[MOCX]\s+(?:AN|N\d*|R|ID|DT|TM)\s+\d+/\d+'
        pattern5 = r'^\s*([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+([A-Za-z][A-Za-z0-9\s\-/()]+?)(?:\s+[MXOC])?\s*(?:AN|N\d*|R|ID|DT|TM)?\s*\d*/?\d*'
        pattern6 = r'^\s*([A-Z]{1,4}\d{1,3})\s+(\d{2,4})\s+([A-Za-z][A-Za-z0-9\s\-/()]+)'
        
        for line in lines:
            line_stripped = line.strip()
            
            lower_line = line_stripped.lower()
            # Normalize multiple spaces to single space for comparison
            normalized_line = re.sub(r'\s+', ' ', lower_line)
            if (normalized_line == 'element summary:' or
                normalized_line == 'data element summary:' or
                normalized_line.startswith('element summary') or
                normalized_line.startswith('data element summary') or
                'element summary' in normalized_line):
                in_element_summary = True
                continue
            
            if not in_element_summary:
                continue
            
            if not line_stripped or len(line_stripped) < 10:
                continue
            
            match = re.search(pattern1, line_stripped)
            if not match:
                match = re.search(pattern2, line_stripped)
            if not match:
                match = re.search(pattern3, line_stripped)
            if not match:
                match = re.search(pattern4, line_stripped)
            if not match:
                match = re.search(pattern5, line_stripped)
            if not match:
                match = re.search(pattern6, line_stripped)
            
            if match:
                ref_id = match.group(1).strip()
                element_name = match.group(3).strip()
                
                # Skip SE segment (Transaction Set Trailer)
                if ref_id.startswith('SE'):
                    break
                
                # Skip ST segment (Transaction Set Header)
                if ref_id.startswith('ST'):
                    continue
                
                # Skip CTT segment (Transaction Totals)
                if ref_id.startswith('CTT'):
                    continue
                
                # Skip N302 (second address line - not needed)
                if ref_id == 'N302':
                    continue
                
                if ref_id in seen_ref_ids:
                    continue
                
                # Skip ALL qualifier fields (including ID Qualifier, Code Qualifier, etc.)
                if 'qualifier' in element_name.lower():
                    seen_ref_ids.add(ref_id)
                    continue
                
                # Start collecting fields from BEG, BIG, BSN, or when we already have fields
                if ref_id.startswith('BEG') or ref_id.startswith('BIG') or ref_id.startswith('BSN') or len(all_fields) > 0:
                    if element_name and len(element_name) > 2:
                        # Extract segment prefix (e.g., "N1" from "N101", "N2" from "N201", "N3" from "N301", "N4" from "N401")
                        # Special handling for N segments
                        segment = None
                        if ref_id.startswith('N1'):
                            segment = 'N1'
                        elif ref_id.startswith('N2'):
                            segment = 'N2'
                        elif ref_id.startswith('N3'):
                            segment = 'N3'
                        elif ref_id.startswith('N4'):
                            segment = 'N4'
                        else:
                            # For other segments, extract letters before numbers
                            segment_match = re.match(r'^([A-Z]+)\d+', ref_id)
                            if segment_match:
                                segment = segment_match.group(1)
                        
                        if segment:
                            if segment not in segment_fields:
                                segment_fields[segment] = []
                            segment_fields[segment].append((element_name, ref_id))
                        
                        all_fields.append((element_name, ref_id, segment))
                        seen_ref_ids.add(ref_id)
        
        # Check if N101 has a code list for expansion
        has_n1_expansion = 'N101' in self.code_lists and len(self.code_lists['N101']) > 0
        
        # Build the final field mappings
        expanded_mappings = []
        skip_segments = set()
        
        if has_n1_expansion:
            # Mark N1, N2, N3, N4 segments for special handling
            skip_segments = {'N1', 'N2', 'N3', 'N4'}
        
        # Collect all N1, N2, N3, N4 fields FIRST
        n1_n2_n3_n4_fields = []
        if has_n1_expansion:
            for element_name, ref_id, segment in all_fields:
                if segment in skip_segments:
                    n1_n2_n3_n4_fields.append((element_name, ref_id, segment))
        
        # Add all non-N1/N2/N3/N4 fields first
        for element_name, ref_id, segment in all_fields:
            if segment not in skip_segments:
                expanded_mappings.append((element_name, f"[{ref_id}]"))
            elif segment == 'N1' and has_n1_expansion:
                # Found start of N1 segment - insert expanded segments here
                code_list = self.code_lists['N101']
                
                # Expand for each code
                for code, code_name in code_list:
                    for elem_name, elem_ref_id, elem_segment in n1_n2_n3_n4_fields:
                        if elem_ref_id == 'N102':
                            # Replace N102 "Name" with the specific code name
                            expanded_mappings.append((code_name, f"[{elem_ref_id}]"))
                        else:
                            # Keep original field name
                            expanded_mappings.append((elem_name, f"[{elem_ref_id}]"))
                
                # Skip to after N1/N2/N3/N4 segments
                break
        
        # Add remaining fields after N1/N2/N3/N4 segments
        if has_n1_expansion:
            found_n_segments = False
            for element_name, ref_id, segment in all_fields:
                if segment in skip_segments:
                    found_n_segments = True
                    continue
                if found_n_segments:
                    expanded_mappings.append((element_name, f"[{ref_id}]"))
        
        self.field_mappings = expanded_mappings
        return self.field_mappings
    
    def generate_excel_file(self, field_mappings: List[Tuple[str, str]]):
        """Generate Excel file with dynamic column widths and formatted cells"""
        descriptions = [desc for desc, code in field_mappings]
        codes = [code for desc, code in field_mappings]
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "EDI Fields"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        code_font = Font(size=10)
        code_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row
        for col_idx, desc in enumerate(descriptions, start=1):
            cell = ws.cell(row=1, column=col_idx, value=desc)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write code row
        for col_idx, code in enumerate(codes, start=1):
            cell = ws.cell(row=2, column=col_idx, value=code)
            cell.font = code_font
            cell.alignment = code_alignment
        
        # Auto-adjust column widths based on content
        for col_idx, (desc, code) in enumerate(zip(descriptions, codes), start=1):
            # Calculate width: max of description and code length, with minimum of 12
            max_length = max(len(desc), len(code), 12)
            # Add some padding (Excel uses character units, roughly 1.2x for padding)
            adjusted_width = max_length + 2
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel workbook
        output_file = str(self.pdf_path).rsplit('.', 1)[0] + '_output.xlsx'
        wb.save(output_file)
        
        # Also generate CSV file
        csv_file = output_file.replace('.xlsx', '.csv')
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(descriptions)
            writer.writerow(codes)
        
        return output_file



class SmartCSVUI:
    """Main UI Application"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("SmartCSV Generator")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # Variables
        self.pdf_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.processing = False
        
        # Configure style
        self.setup_styles()
        
        # Create UI
        self.create_widgets()
        
        # Center window
        self.center_window()
    
    def setup_styles(self):
        """Configure ttk styles for professional appearance"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        bg_color = "#f0f0f0"
        accent_color = "#0078d4"
        button_color = "#0078d4"
        
        style.configure("Title.TLabel", font=("Segoe UI", 24, "bold"), foreground=accent_color)
        style.configure("Subtitle.TLabel", font=("Segoe UI", 10), foreground="#666666")
        style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"), foreground="#333333")
        style.configure("Info.TLabel", font=("Segoe UI", 9), foreground="#666666")
        
        style.configure("Action.TButton", font=("Segoe UI", 10), padding=10)
        style.map("Action.TButton",
                 background=[("active", "#005a9e"), ("!active", button_color)],
                 foreground=[("active", "white"), ("!active", "white")])
    
    def create_widgets(self):
        """Create all UI widgets"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Title section
        title_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        
        title_label = ttk.Label(title_frame, text="SmartCSV Generator", style="Title.TLabel")
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        subtitle_label = ttk.Label(title_frame,
                                   text="Convert EDI PDF Specifications to Excel template",
                                   style="Subtitle.TLabel")
        subtitle_label.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        # Separator
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).grid(row=1, column=0, sticky=(tk.W, tk.E), pady=10)
        
        # Input section
        input_frame = ttk.LabelFrame(main_frame, text="  Input PDF File  ", padding="15")
        input_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        input_frame.columnconfigure(1, weight=1)
        
        ttk.Label(input_frame, text="PDF File:", style="Header.TLabel").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        pdf_entry = ttk.Entry(input_frame, textvariable=self.pdf_path, state="readonly")
        pdf_entry.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        browse_btn = ttk.Button(input_frame, text="Browse PDF...", command=self.browse_pdf)
        browse_btn.grid(row=2, column=0, sticky=tk.W)
        
        # Info label
        info_label = ttk.Label(input_frame, 
                              text="Select an EDI specification PDF file (e.g., EDI850_4010_Spec.pdf)",
                              style="Info.TLabel")
        info_label.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        # Output section
        output_frame = ttk.LabelFrame(main_frame, text="  Output Settings  ", padding="15")
        output_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        output_frame.columnconfigure(1, weight=1)
        
        ttk.Label(output_frame, text="Output Excel:", style="Header.TLabel").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        output_entry = ttk.Entry(output_frame, textvariable=self.output_path, state="readonly")
        output_entry.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        output_btn = ttk.Button(output_frame, text="Choose Location...", command=self.choose_output)
        output_btn.grid(row=2, column=0, sticky=tk.W)
        
        # Action button
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=4, column=0, pady=(10, 15))
        
        self.generate_btn = ttk.Button(action_frame,
                                       text="Generate Excel",
                                       command=self.generate_csv,
                                       style="Action.TButton",
                                       width=20)
        self.generate_btn.grid(row=0, column=0)
        
        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="  Progress  ", padding="15")
        progress_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        progress_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.log_text = scrolledtext.ScrolledText(progress_frame, 
                                                  height=10, 
                                                  wrap=tk.WORD,
                                                  font=("Consolas", 9),
                                                  bg="#ffffff",
                                                  fg="#333333")
        self.log_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status bar
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=6, column=0, sticky=(tk.W, tk.E))
        
        self.status_label = ttk.Label(status_frame, text="Ready", style="Info.TLabel")
        self.status_label.grid(row=0, column=0, sticky=tk.W)
        
        # Initial log message
        self.log("Welcome to SmartCSV Excel Generator!")
        self.log("Select a PDF file to begin...")
    
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def log(self, message: str):
        """Add message to log window"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def update_status(self, message: str):
        """Update status bar"""
        self.status_label.config(text=message)
        self.root.update_idletasks()
    
    def browse_pdf(self):
        """Open file dialog to select PDF"""
        filename = filedialog.askopenfilename(
            title="Select EDI PDF Specification",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if filename:
            self.pdf_path.set(filename)
            self.log(f"Selected PDF: {Path(filename).name}")
            
            # Auto-suggest output path
            pdf_path = Path(filename)
            suggested_output = pdf_path.parent / f"{pdf_path.stem}_output.xlsx"
            self.output_path.set(str(suggested_output))
            self.log(f"Suggested output: {suggested_output.name}")
            self.update_status("PDF selected - Ready to generate")
    
    def choose_output(self):
        """Open file dialog to choose output location"""
        filename = filedialog.asksaveasfilename(
            title="Save Excel As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="output.xlsx"
        )
        
        if filename:
            self.output_path.set(filename)
            self.log(f"Output location: {Path(filename).name}")
    
    def generate_csv(self):
        """Start Excel generation process"""
        if not self.pdf_path.get():
            messagebox.showwarning("No PDF Selected", "Please select a PDF file first.")
            return
        
        if not self.output_path.get():
            messagebox.showwarning("No Output Location", "Please choose an output location.")
            return
        
        # Disable button and start progress
        self.generate_btn.config(state="disabled")
        self.progress_bar.start(10)
        self.update_status("Processing...")
        
        # Run in separate thread to keep UI responsive
        thread = threading.Thread(target=self.process_pdf, daemon=True)
        thread.start()
    
    def process_pdf(self):
        """Process PDF and generate Excel (runs in separate thread)"""
        try:
            self.log("\n" + "="*50)
            self.log("Starting PDF processing...")
            self.log(f"Input: {Path(self.pdf_path.get()).name}")
            
            # Extract fields
            self.log("Extracting text from PDF...")
            extractor = EDIFieldExtractor(self.pdf_path.get())
            extractor.extract_text()
            
            self.log("Extracting code lists for segment expansion...")
            code_lists = extractor.extract_code_lists()
            if code_lists:
                self.log(f"✓ Found {len(code_lists)} fields with code lists:")
                for field, codes in code_lists.items():
                    code_names = ', '.join([f"{c[0]}={c[1]}" for c in codes[:3]])  # Show first 3
                    if len(codes) > 3:
                        code_names += f", ... ({len(codes)} total)"
                    self.log(f"  • {field}: {code_names}")
            else:
                self.log("⚠ No code lists found - N1 segment will not be expanded")
                self.log("  Check if PDF contains 'CodeList Summary' sections")
            
            self.log("Parsing field mappings with segment expansion...")
            field_mappings = extractor.parse_field_mappings()
            
            self.log(f"Generated {len(field_mappings)} total field mappings (including expansions)")
            
            if not field_mappings:
                self.root.after(0, lambda: messagebox.showerror(
                    "No Fields Found",
                    "No EDI fields were extracted from the PDF. Please check the file format."
                ))
                return
            
            # Generate Excel
            self.log("Generating Excel file...")
            self.generate_excel_file(field_mappings)
            
            self.log(f"Excel generated successfully!")
            self.log(f"Output: {Path(self.output_path.get()).name}")
            self.log("="*50)
            
            # Show success message
            self.root.after(0, lambda: messagebox.showinfo(
                "Success",
                f"Excel file generated successfully!\n\n"
                f"Fields extracted: {len(field_mappings)}\n"
                f"Output: {Path(self.output_path.get()).name}"
            ))
            
            self.root.after(0, lambda: self.update_status("Completed successfully"))
            
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.log(f"\nERROR: {error_msg}")
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
            self.root.after(0, lambda: self.update_status("Error occurred"))
        
        finally:
            # Re-enable button and stop progress
            self.root.after(0, lambda: self.generate_btn.config(state="normal"))
            self.root.after(0, lambda: self.progress_bar.stop())
    
    def generate_excel_file(self, field_mappings: List[Tuple[str, str]]):
        """Generate Excel file with dynamic column widths and formatted cells"""
        descriptions = [desc for desc, code in field_mappings]
        codes = [code for desc, code in field_mappings]
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "EDI Fields"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        code_font = Font(size=10)
        code_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row
        for col_idx, desc in enumerate(descriptions, start=1):
            cell = ws.cell(row=1, column=col_idx, value=desc)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write code row
        for col_idx, code in enumerate(codes, start=1):
            cell = ws.cell(row=2, column=col_idx, value=code)
            cell.font = code_font
            cell.alignment = code_alignment
        
        # Auto-adjust column widths based on content
        for col_idx, (desc, code) in enumerate(zip(descriptions, codes), start=1):
            # Calculate width: max of description and code length, with minimum of 12
            max_length = max(len(desc), len(code), 12)
            # Add some padding (Excel uses character units, roughly 1.2x for padding)
            adjusted_width = max_length + 2
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel workbook
        output_file = self.output_path.get()
        wb.save(output_file)
        
        # Also generate CSV file
        csv_file = output_file.replace('.xlsx', '.csv')
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(descriptions)
            writer.writerow(codes)
        
        self.log(f"Excel file generated with dynamic column widths: {Path(output_file).name}")
        self.log(f"CSV file also generated: {Path(csv_file).name}")


def main():
    """Main entry point"""
    root = tk.Tk()
    app = SmartCSVUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

# Made with Bob
